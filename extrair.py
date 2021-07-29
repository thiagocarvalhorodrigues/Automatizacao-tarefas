from selenium import webdriver
from selenium.webdriver.chrome.options import Options
import os
import time
import openpyxl
from selenium.webdriver.common.by import By

class WebScraping:
    def __init__(self):
        chrome_options = Options()
        chrome_options.add_argument('--lang=pt-BR')
        chrome_options.add_argument("--start-maximized")
        self.driver = webdriver.Chrome(executable_path=os.getcwd() + os.sep + 'driver' +  os.sep +
                                                       'chromedriver.exe', options=chrome_options)

    def Iniciar(self):
        self.cria_cabecalho_planilha()
        self.acessar_site()


    def cria_cabecalho_planilha(self):
        self.planilha = openpyxl.Workbook()
        self.planilha.create_sheet('Oportunidades')
        self.planilha_vagas_emprego_cadmus = self.planilha['Oportunidades']
        self.planilha_vagas_emprego_cadmus.cell(row=1, column=1, value='Nome')
        self.planilha_vagas_emprego_cadmus.cell(row=1, column=2, value='Local')


    def acessar_site(self):
        self.driver.get("https://cadmus.com.br/vagas-tecnologia/")
        self.captura_informacoes()

    def captura_informacoes(self):
        xpath = '//section[@class="solucoes-texto"]'
        elemento = self.driver.find_element(By.XPATH, xpath )
        self.driver.execute_script("return arguments[0].scrollIntoView();", elemento)
        time.sleep(5)
        vagas = self.driver.find_elements(By.XPATH, '//*[@class="box"]//h3')
        local = self.driver.find_elements(By.XPATH, '//*[@class="container pt-4 pb-4"]//*[@id="pfolio"]//*[@class="local"]')
        desc = self.driver.find_elements(By.XPATH, '//*[@class="btn azul"]')

        '//*[@class="box z-depth-1"]//p'
        self.gravar_informacoes(vagas, local)

    def gravar_informacoes(self, vagas, local):
        for item in range(0, len(vagas)):
            self.lista = []
            dados = [
                vagas[item].text,
                local[item].text,
            ]
            self.planilha_vagas_emprego_cadmus.append(dados)
            self.planilha.save('vagas.xlsx')
            self.lista.append(dados)


            print(self.lista)
        print("||||A sua tabela já está pronta||||")
        print("||||Estamos anexando e enviando para o seu destinatário.||||")
        self.driver.close()
        self.driver.quit()